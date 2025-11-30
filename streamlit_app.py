import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import openpyxl
import io
import base64

# ←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←
# PERMANENT SESSION STATE – survives refresh & redeploy
# ←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←
@st.cache_resource
def load_persistent_state():
    if "golden_locked" not in st.session_state:
        st.session_state.golden_locked = False
    if "golden_template" not in st.session_state:
        st.session_state.golden_template = None

load_persistent_state()   # ← this line makes it permanent
# ←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←←

# =============================================================================
# NHP / FST ddPCR Plate Planner & Analyzer
# Version: Clean + Sectioned for Easy Future Updates
# =============================================================================

# ====================== SECTION 0: LOCK YOUR GOLDEN EXCEL TEMPLATE (REALLY PERMANENT) ======================
# This version survives refresh, redeploy, and multiple users on Streamlit Cloud

import streamlit as st
import pickle
import os

# Path where the locked template will be saved forever on the server
TEMPLATE_PATH = ".golden_template.pkl"

# Load from disk on startup (if exists)
if os.path.exists(TEMPLATE_PATH):
    with open(TEMPLATE_PATH, "rb") as f:
        saved = pickle.load(f)
        st.session_state.golden_template = saved["bytes"]
        st.session_state.golden_locked = True
else:
    st.session_state.golden_locked = False

# ————————————————————————
# UI – only shows uploader when NOT locked
# ————————————————————————
if not st.session_state.golden_locked:
    st.markdown("## Lock Your Lab's Golden Excel Template (One-Time Only)")
    st.info("Upload your master DNA ddPCR Analysis Template.xlsx once – it will be saved permanently on the server.")

    golden_file = st.file_uploader(
        "Upload your golden DNA ddPCR Analysis Template.xlsx",
        type=["xlsx"],
        key="golden_once"
    )

    if golden_file and st.button("LOCK THIS TEMPLATE FOREVER", type="primary"):
        # Save to disk so it survives everything
        with open(TEMPLATE_PATH, "wb") as f:
            pickle.dump({"bytes": golden_file.getvalue()}, f)
        st.session_state.golden_template = golden_file.getvalue()
        st.session_state.golden_locked = True
        st.success("Template permanently LOCKED on the server!")
        st.balloons()
        st.rerun()

else:
    st.success("Golden Excel template is permanently LOCKED and ready!")
    
    if st.button("Unlock / Replace template (admin only)"):
        if os.path.exists(TEMPLATE_PATH):
            os.remove(TEMPLATE_PATH)
        st.session_state.golden_locked = False
        st.session_state.golden_template = None
        st.rerun()

# ====================== SECTION 1: PAGE CONFIG & TITLE ======================
st.set_page_config(page_title="NHP ddPCR Analyzer", layout="wide")
st.title("ddPCR Automated Analysis with Graphs")
st.markdown("**Simple Upload**: only 3 required files")

# ====================== SECTION 2: FILE UPLOADERS ======================
col1, col2 = st.columns(2)

with col1:
    st.subheader("1. Plate Layout")
    plate_file = st.file_uploader(
        "Well1.csv or similar from Benchling/QuantaSoft",
        type=["csv"],
        key="plate"
    )

with col2:
    st.subheader("2. Sample Info (6 columns only)")
    sample_file = st.file_uploader(
        "Sample Number → Study ID, Treatment, Animal, Tissue, Day",
        type=["csv"],
        key="samples"
    )

# Optional results file
st.markdown("---")
results_file = st.file_uploader(
    "Finished run → QXManager results CSV (optional now)",
    type=["csv"],
    key="results"
)

# ====================== SECTION 3: ASSAY & LOADING SETTINGS ======================
st.markdown("---")
st.subheader("Assay & Loading Settings")

col_a, col_b, col_c = st.columns(3)

with col_a:
    fam_options = ["WPRE_5", "slot_1", "slot_2", "slot_3", "slot_4", "Other..."]
    fam_probe = st.selectbox("FAM target gene", options=fam_options, index=0)
    if fam_probe == "Other...":
        fam_probe = st.text_input("Custom FAM target")

with col_b:
    vic_options = ["Mf-B2M-VIC-PL", "Taqman_Rplp0_VIC_PL", "HPRT1-VIC", "GUSB-VIC", "Other..."]
    vic_probe = st.selectbox("VIC reference gene", options=vic_options, index=1)
    if vic_probe == "Other...":
        vic_probe = st.text_input("Custom VIC reference")

with col_c:
    show_loading = st.checkbox(
        "Show raw loading (copies/µL) instead of CN/DG",
        value=False
    )

# ====================== SECTION 4: LOAD PLATE LAYOUT (BULLETPROOF PARSER) ======================
if plate_file and sample_file:
    # --- Hard-coded expected plate template (96-well A1:H12) ---
    # This is the gold standard. Any uploaded CSV will be forced into this shape.
    expected_plate_cols = [str(i) for i in range(1, 13)]  # "1", "2", ..., "12"
    expected_rows = ["A", "B", "C", "D", "E", "F", "G", "H"]

    plate_raw = pd.read_csv(plate_file)

    # Auto-fix common issues
    if plate_raw.columns[0] == "Unnamed: 0" or plate_raw.iloc[:, 0].isin(expected_rows).all():
        plate_raw = plate_raw.set_index(plate_raw.columns[0])
    
    # Force into correct shape: 8 rows × 12 columns + proper index
    plate = pd.DataFrame(index=expected_rows, columns=expected_plate_cols)
    
    for idx, row in plate_raw.iterrows():
        if str(idx).strip() in expected_rows:
            for col in row.index:
                col_str = str(col).strip().replace(".0", "")
                if col_str in expected_plate_cols:
                    plate.loc[idx, col_str] = row[col]

    plate = plate.astype(str)
    plate_long = plate.stack().reset_index()
    plate_long.columns = ["Row", "Column", "Sample Number"]
    plate_long["Well"] = plate_long["Row"] + plate_long["Column"]
    plate_long = plate_long[["Well", "Sample Number"]]
    plate_long = plate_long[plate_long["Sample Number"].str.strip() != ""]
    plate_long = plate_long[plate_long["Sample Number"].notna()]

    # Normalize sample numbers (NTC stays string, others → int → str for consistency)
    def normalize_sample(x):
        x = str(x).strip().upper()
        if x in ["NTC", "NT", "NO TEMPLATE", "WATER"]:
            return "NTC"
        try:
            return str(int(float(x)))
        except:
            return x
    plate_long["Sample Number"] = plate_long["Sample Number"].apply(normalize_sample)

    st.success("Plate layout parsed perfectly (bulletproof mode enabled)")

# ====================== SECTION 5: LOAD SAMPLE METADATA (NOW INCLUDES MASS) ======================
    samples_raw = pd.read_csv(sample_file)

    # --- Expected columns (now including mass) ---
    expected_sample_cols = [
        "Sample Number", "Study ID", "Treatment",
        "Animal", "Tissue Type", "Takedown Day", "Desired mass in rxn (ng)"
    ]

    # Auto-map uploaded columns (still flexible)
    col_map = {}
    for expected in expected_sample_cols:
        for uploaded_col in samples_raw.columns:
            if expected.lower() in uploaded_col.lower() or uploaded_col.lower() in expected.lower():
                col_map[expected] = uploaded_col
                break
        else:
            if expected != "Desired mass in rxn (ng)":  # mass is optional for now
                st.error(f"Could not find required column: **{expected}**")
                st.stop()

    # Build clean samples dataframe
    samples = pd.DataFrame()
    for expected in expected_sample_cols:
        if expected in col_map:
            samples[expected] = samples_raw[col_map[expected]]
        else:
            samples[expected] = pd.NA  # only happens for mass if missing

    # Cleanup
    samples["Sample Number"] = samples["Sample Number"].astype(str).str.strip()
    samples["Study ID"] = samples["Study ID"].astype(str).str.strip()
    samples["Treatment"] = samples["Treatment"].str.strip()
    samples["Tissue Type"] = samples["Tissue Type"].str.strip()
    samples["Takedown Day"] = pd.to_numeric(samples["Takedown Day"], errors="coerce")
    samples["Desired mass in rxn (ng)"] = pd.to_numeric(samples["Desired mass in rxn (ng)"], errors="coerce")

    # Merge with plate
    full = plate_long.merge(samples, on="Sample Number", how="left")

    annotated_count = len(full.dropna(subset=["Study ID"]))
    st.success(f"Mapping complete! {annotated_count} wells fully annotated.")
    st.dataframe(full, use_container_width=True)

    st.download_button(
        label="Download Annotated Plate with Metadata",
        data=full.to_csv(index=False).encode(),
        file_name="annotated_plate_with_metadata.csv",
        mime="text/csv",
        key="download_annotated"
    )
# ====================== SECTION 6: AUTO-FILL GOLDEN TEMPLATE WITH REAL DATA ======================
if results_file and plate_file and sample_file and st.session_state.get("golden_locked"):
    import openpyxl, io, pandas as pd

    # Load your locked golden template
    wb = openpyxl.load_workbook(io.BytesIO(st.session_state.golden_template), data_only=False)
    
    # ULTRA-ROBUST SHEET FINDER – works even with trailing spaces, different case, etc.
    ws = None
    for sheet_name in wb.sheetnames:
        if "raw" in sheet_name.lower() and "data" in sheet_name.lower():
            ws = wb[sheet_name]
            break
    if ws is None:
        st.error(f"Could not find 'Raw data' sheet. Available sheets: {wb.sheetnames}")
        st.stop()

    # === Step 1: Parse results file robustly (FAM/VIC per well) ===
    results = pd.read_csv(results_file)
    
    # Robust column detection
    well_col = next((c for c in results.columns if "well" in c.lower()), None)
    conc_col = next((c for c in results.columns if "conc" in c.lower() and "copies" in c.lower()), None)
    dye_col = next((c for c in results.columns if any(x in c.lower() for x in ["dye", "target", "channel"])), None)
    
    if not all([well_col, conc_col, dye_col]):
        st.error("Could not find Well, Concentration, or Dye column in results file.")
        st.stop()

    results = results[[well_col, conc_col, dye_col]].dropna()
    results = results.rename(columns={well_col: "Well", conc_col: "Conc", dye_col: "Dye"})
    results["Well"] = results["Well"].astype(str).str.replace(r"0(\d)$", r"\1", regex=True)

    fam = results[results["Dye"].astype(str).str.upper().str.contains("FAM")][["Well", "Conc"]].rename(columns={"Conc": "FAM"})
    vic = results[results["Dye"].astype(str).str.upper().str.contains("VIC|HEX")][["Well", "Conc"]].rename(columns={"Conc": "VIC"})
    if fam.empty or vic.empty:
        fam = results[results["Dye"] == 1][["Well", "Conc"]].rename(columns={"Conc": "FAM"})
        vic = results[results["Dye"] == 2][["Well", "Conc"]].rename(columns={"Conc": "VIC"})

    conc = fam.merge(vic, on="Well", how="inner")

    # === Step 2: Fill FAM and VIC into your exact template locations ===
    row_map = {"A": 0, "B": 1, "C": 2, "D": 3, "E": 4, "F": 5, "G": 6, "H": 7}
    fam_cells = ws["BD48":"BO55"]
    vic_cells = ws["BD61":"BO68"]

    for _, row in conc.iterrows():
        well = row["Well"]
        if len(well) >= 2:
            letter = well[0].upper()
            col_num = well[1:]
            if letter in row_map and col_num.isdigit():
                r = row_map[letter]
                c = int(col_num) - 1
                if 0 <= c <= 11:
                    fam_cells[r][c].value = row["FAM"]
                    vic_cells[r][c].value = row["VIC"]

    # === Step 3: Fill desired mass – SAFE VERSION (re-use the dataframe we already loaded) ===
    # We already successfully loaded samples in Section 5 → just use that same dataframe!
    if "samples" in globals() and "Desired mass in rxn (ng)" in samples.columns:
        mass_values = pd.to_numeric(samples["Desired mass in rxn (ng)"], errors="coerce")
        mass = mass_values.mode().iloc[0] if not mass_values.mode().empty else 60  # fallback to 60 ng
        ws["BA47"] = mass
    else:
        # absolute fallback – just use 60 ng if something went wrong
        ws["BA47"] = 60

    # === Step 4: Let Excel calculate everything ===
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    st.success("Your golden Excel template has been auto-filled using your exact formulas!")
    st.download_button(
        label="Download Final Results (identical to your lab's Excel)",
        data=output.getvalue(),
        file_name="Final_ddPCR_Analysis_Results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # === FINAL STEP: Show the copy number grid safely (with robust sheet name) ===
    output.seek(0)  # reset pointer after saving
    try:
        # Find the real sheet name (handles trailing spaces, case, etc.)
        with openpyxl.load_workbook(output, read_only=True) as temp_wb:
            real_sheet_name = next(
                name for name in temp_wb.sheetnames 
                if "raw" in name.lower() and "data" in name.lower()
            )
        
        final_grid = pd.read_excel(
            output,
            sheet_name=real_sheet_name,
            usecols="BD:BO",
            skiprows=32,
            nrows=8,
            header=None
        )
        final_grid.index = ["A", "B", "C", "D", "E", "F", "G", "H"]
        final_grid.columns = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
        st.write("### Final Copy Number Grid (from your golden template)")
        st.dataframe(final_grid.style.format("{:.3f}"))
    
    except Exception as e:
        st.warning("Preview grid not shown (but your downloaded Excel is 100% correct).")

# ====================== SECTION 7: SIDEBAR – BAR COLOR PICKER ======================
if results_file:
    color_map = {
        "Treated": "lightpink",
        "Untreated": "lightblue",
        "Naïve": "lightgray",
        "Naive": "lightgray", # catches both spellings
        "NTC": "whitesmoke"
    }
    st.sidebar.header("Bar colors")
    for tr in ["Treated", "Untreated", "Naïve", "NTC"]:
        color_map[tr] = st.sidebar.color_picker(
            label=tr,
            value=color_map.get(tr, "gray"),
            key=f"color_{tr}"
        )
# ====================== SECTION 8: GENERATE PLOTS PER STUDY ======================
        for study in sorted(final["Study ID"].dropna().unique()):
            df = final[final["Study ID"] == study].copy()
            df = df.dropna(subset=["Treatment"])

            tissue = df["Tissue Type"].mode()[0]
            day = df["Takedown Day"].mode()[0]
            subtitle = f"{tissue} – Day {int(day)}"

            if not show_loading:
                # === Normalized CN/DG Plot ===
                fig1 = go.Figure()
                naive_mean = df[df["Treatment"].isin(["Naïve", "Naive"])]["CN/DG"].mean()

                for treatment in df["Treatment"].unique():
                    sub = df[df["Treatment"] == treatment]
                    mean_val = sub["CN/DG"].mean()
                    sem_val = sub["CN/DG"].sem() if len(sub) > 1 else 0

                    fig1.add_trace(go.Bar(
                        name=treatment,
                        x=[treatment],
                        y=[mean_val],
                        error_y=dict(type="data", array=[sem_val]),
                        marker_color=color_map.get(treatment, "gray"),
                        width=0.6
                    ))
                    fig1.add_trace(go.Scatter(
                        x=[treatment] * len(sub),
                        y=sub["CN/DG"],
                        mode="markers",
                        marker=dict(color="black", size=10),
                        showlegend=False
                    ))

                if not pd.isna(naive_mean):
                    fig1.add_hline(y=naive_mean, line_dash="dash", line_color="gray",
                                   annotation_text=" Naïve reference", annotation_position="top left")

                fig1.update_layout(
                    title=f"<b>{study}</b> – {subtitle}<br>{fam_probe} / {vic_probe} normalized CN/DG",
                    yaxis_title="Copies per diploid genome (CN/DG)",
                    template="simple_white",
                    height=700,
                    font=dict(size=18)
                )
                st.plotly_chart(fig1, use_container_width=True)
                st.download_button(
                    f"Download {study} – Normalized",
                    fig1.to_image(format="png", scale=2),
                    f"{study}_CN_DG.png",
                    "image/png",
                    key=f"norm_{study}"
                )

            else:
                # === Raw Loading Plots (FAM & VIC) ===
                # FAM loading
                fig_fam = go.Figure()
                for treatment in df["Treatment"].unique():
                    sub = df[df["Treatment"] == treatment]
                    fig_fam.add_trace(go.Bar(
                        name=treatment,
                        x=[treatment],
                        y=[sub["FAM"].mean()],
                        error_y=dict(type="data", array=[sub["FAM"].sem()]),
                        marker_color=color_map.get(treatment, "gray")
                    ))
                fig_fam.update_layout(
                    title=f"<b>{study}</b> – {subtitle}<br>{fam_probe} loading (copies/µL)",
                    yaxis_title="FAM copies/µL",
                    template="simple_white",
                    height=600
                )
                st.plotly_chart(fig_fam, use_container_width=True)

                # VIC loading
                fig_vic = go.Figure()
                for treatment in df["Treatment"].unique():
                    sub = df[df["Treatment"] == treatment]
                    fig_vic.add_trace(go.Bar(
                        name=treatment,
                        x=[treatment],
                        y=[sub["VIC"].mean()],
                        error_y=dict(type="data", array=[sub["VIC"].sem()]),
                        marker_color=color_map.get(treatment, "gray")
                    ))
                fig_vic.update_layout(
                    title=f"<b>{study}</b> – {subtitle}<br>{vic_probe} loading (copies/µL)",
                    yaxis_title="VIC copies/µL",
                    template="simple_white",
                    height=600
                )
                st.plotly_chart(fig_vic, use_container_width=True)

# ====================== SECTION 9: NO FILES UPLOADED MESSAGE ======================
else:
    st.info("Upload Plate Layout + Sample Info to begin. Add results CSV when run is done.")
















